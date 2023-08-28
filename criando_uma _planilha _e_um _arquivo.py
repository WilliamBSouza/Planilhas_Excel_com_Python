import openpyxl

#criar uma planilha em excel (livro)
livro = openpyxl.Workbook()
#como visualizar as páginas existentes 
print(livro.sheetnames)
#como criar uma página 
livro.create_sheet('Celulares')
#como selecionar uma página 
Celulares_page = livro['Celulares']
Celulares_page.append(['Aparelho', 'Quantidade','Valor'])
Celulares_page.append(['Iphone 14 Pro Max', '5' , 'R$10.000,00'])
Celulares_page.append(['Iphone 14 Pro', '2 ' , 'R$7.000,00'])
Celulares_page.append(['Iphone 14', '10' , 'R$5.000,00'])
Celulares_page.append(['Iphone 13 Pro', '7' , 'R$6.000,00'])
Celulares_page.append(['Iphone 13 Pro Max', '6' , 'R$6.500,00'])
#salvar a planilha 
livro.save('Livro de Celulares.xlsx')
