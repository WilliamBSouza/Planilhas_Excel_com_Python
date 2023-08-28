import openpyxl

#abrir o arquivo 
livro = openpyxl.load_workbook('livro de celulares.xlsx')
#selecionando a pagina 
celulare_page = livro['Celulares']
#imprimindo os dados de cada linha a partir da linha 2 e parando na linha 5 
for linhas in celulare_page.iter_rows(min_row=2,max_row=5):
    for celula in linhas:
        print(celula.value)
print(200*'-')

#imprimindo os dados de cada linha como uma linha e sem limitações de linha

for linhas in celulare_page.iter_rows():
        print(linhas[0].value,linhas[1].value,linhas[2].value)
print(200*'-')


#imprimindo os dados de cada linha como uma linha e sem limitações de linha e separando os dados com um traço e com espaço para melhor separa-los

for linhas in celulare_page.iter_rows():
        print(f'{linhas[0].value}   -     {linhas[1].value}  -       {linhas[2].value}')

print(200*'-')


#para modificar um item na planilha através do que está escrito na celula 

for linhas in celulare_page.iter_rows():
      for linha in linhas:
            if linha.value =="Iphone 14 Pro":
                  linha.value = 'Iphone 14 Pro 256gb'
livro.save('Livro de Celulares.xlsx') # para separar a altração é só colocar v2 ou a versão que desejar para a alteração

